"""XLSX-backed scheme storage for grid-planning parameter maintenance."""

from __future__ import annotations

import re
import shutil
from copy import deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


WEB_ROOT = Path(__file__).resolve().parent
DEFAULT_SCHEME_ROOT = WEB_ROOT / "planning_schemes"
WORKBOOK_NAME = "parameters.xlsx"

SHEET_SPECS: dict[str, tuple[str, list[str]]] = {
    "time_series": ("8760时序数据", ["hour_index", "datetime", "wind_speed", "solar_irradiance", "load"]),
    "diesel_generators": (
        "柴发参数",
        [
            "name",
            "capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "power_upper",
            "power_lower",
            "fuel_rate",
        ],
    ),
    "wind_turbines": (
        "风机参数",
        [
            "name",
            "capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "cut_in_wind_speed",
            "cut_out_wind_speed",
        ],
    ),
    "photovoltaics": (
        "光伏参数",
        [
            "name",
            "capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "cut_in_wind_speed",
            "cut_out_wind_speed",
        ],
    ),
    "storage_pcs": (
        "储能PCS参数",
        ["name", "power_capacity", "design_capacity_lower", "design_capacity_upper", "cost"],
    ),
    "storage_battery_packs": (
        "储能电池组参数",
        ["name", "battery_capacity", "design_capacity_lower", "design_capacity_upper", "cost"],
    ),
    "hydrogen_electrolyzers": (
        "电制氢参数",
        [
            "name",
            "power_capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "electric_to_hydrogen_efficiency",
        ],
    ),
    "hydrogen_tanks": (
        "储氢罐参数",
        ["name", "hydrogen_tank_capacity", "design_capacity_lower", "design_capacity_upper", "cost"],
    ),
    "fuel_cells": (
        "燃料电池参数",
        [
            "name",
            "power_capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "hydrogen_to_electric_efficiency",
        ],
    ),
}

DEFAULT_DEVICE_ROWS: dict[str, list[dict[str, Any]]] = {
    "diesel_generators": [
        {
            "name": "柴发1",
            "capacity": 100,
            "design_capacity_lower": 0,
            "design_capacity_upper": 500,
            "cost": 0,
            "power_upper": 100,
            "power_lower": 20,
            "fuel_rate": 0.26,
        }
    ],
    "wind_turbines": [
        {
            "name": "风机1",
            "capacity": 50,
            "design_capacity_lower": 0,
            "design_capacity_upper": 1000,
            "cost": 0,
            "cut_in_wind_speed": 3,
            "cut_out_wind_speed": 25,
        }
    ],
    "photovoltaics": [
        {
            "name": "光伏1",
            "capacity": 50,
            "design_capacity_lower": 0,
            "design_capacity_upper": 1000,
            "cost": 0,
            "cut_in_wind_speed": 0,
            "cut_out_wind_speed": 0,
        }
    ],
    "storage_pcs": [
        {"name": "储能PCS1", "power_capacity": 50, "design_capacity_lower": 0, "design_capacity_upper": 500, "cost": 0}
    ],
    "storage_battery_packs": [
        {
            "name": "储能电池组1",
            "battery_capacity": 200,
            "design_capacity_lower": 0,
            "design_capacity_upper": 2000,
            "cost": 0,
        }
    ],
    "hydrogen_electrolyzers": [
        {
            "name": "电制氢1",
            "power_capacity": 50,
            "design_capacity_lower": 0,
            "design_capacity_upper": 500,
            "cost": 0,
            "electric_to_hydrogen_efficiency": 0.7,
        }
    ],
    "hydrogen_tanks": [
        {
            "name": "储氢罐1",
            "hydrogen_tank_capacity": 100,
            "design_capacity_lower": 0,
            "design_capacity_upper": 2000,
            "cost": 0,
        }
    ],
    "fuel_cells": [
        {
            "name": "燃料电池1",
            "power_capacity": 50,
            "design_capacity_lower": 0,
            "design_capacity_upper": 500,
            "cost": 0,
            "hydrogen_to_electric_efficiency": 0.55,
        }
    ],
}

INVALID_NAME_RE = re.compile(r'[<>:"/\\|?*]')


def validate_scheme_name(name: str) -> str:
    clean = str(name or "").strip()
    if clean in {"", ".", ".."} or INVALID_NAME_RE.search(clean) or ".." in clean:
        raise ValueError("方案名称不能为空，且不能包含路径或非法字符")
    return clean


def default_time_series() -> list[dict[str, Any]]:
    return [
        {"hour_index": hour, "datetime": f"H{hour:04d}", "wind_speed": 0, "solar_irradiance": 0, "load": 0}
        for hour in range(1, 8761)
    ]


def default_payload(scheme: str) -> dict[str, Any]:
    payload: dict[str, Any] = {"scheme": scheme, "time_series": default_time_series(), "validation": []}
    for key in SHEET_SPECS:
        if key != "time_series":
            payload[key] = deepcopy(DEFAULT_DEVICE_ROWS[key])
    payload["capacity_limits"] = []
    return payload


@dataclass
class PlanningStore:
    root: Path = DEFAULT_SCHEME_ROOT

    def __post_init__(self) -> None:
        self.root = Path(self.root).resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def scheme_dir(self, name: str) -> Path:
        clean = validate_scheme_name(name)
        path = (self.root / clean).resolve()
        if self.root not in path.parents and path != self.root:
            raise ValueError("方案路径越界")
        return path

    def workbook_path(self, name: str) -> Path:
        return self.scheme_dir(name) / WORKBOOK_NAME

    def list_schemes(self) -> list[dict[str, Any]]:
        schemes: list[dict[str, Any]] = []
        for folder in sorted(self.root.iterdir(), key=lambda item: item.name):
            if not folder.is_dir():
                continue
            workbook = folder / WORKBOOK_NAME
            schemes.append(
                {
                    "name": folder.name,
                    "has_workbook": workbook.exists(),
                    "modified_at": workbook.stat().st_mtime if workbook.exists() else None,
                }
            )
        return schemes

    def create_scheme(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if folder.exists():
            raise FileExistsError(f"方案已存在: {clean}")
        folder.mkdir(parents=True)
        self.write_scheme(clean, default_payload(clean))
        return self.read_scheme(clean)

    def copy_scheme(self, source: str, target: str) -> dict[str, Any]:
        source_dir = self.scheme_dir(source)
        target_dir = self.scheme_dir(target)
        if not source_dir.exists():
            raise FileNotFoundError(f"源方案不存在: {source}")
        if target_dir.exists():
            raise FileExistsError(f"目标方案已存在: {target}")
        shutil.copytree(source_dir, target_dir)
        return self.read_scheme(target)

    def rename_scheme(self, source: str, target: str) -> dict[str, Any]:
        source_dir = self.scheme_dir(source)
        target_dir = self.scheme_dir(target)
        if not source_dir.exists():
            raise FileNotFoundError(f"源方案不存在: {source}")
        if target_dir.exists():
            raise FileExistsError(f"目标方案已存在: {target}")
        source_dir.rename(target_dir)
        return self.read_scheme(target)

    def write_scheme(self, name: str, payload: dict[str, Any]) -> None:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        folder.mkdir(parents=True, exist_ok=True)
        workbook = build_workbook(payload | {"scheme": clean})
        tmp_path = folder / f".{WORKBOOK_NAME}.tmp"
        final_path = folder / WORKBOOK_NAME
        workbook.save(tmp_path)
        tmp_path.replace(final_path)

    def read_scheme(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        path = self.workbook_path(clean)
        if not path.exists():
            raise FileNotFoundError(f"方案参数文件不存在: {path}")
        payload = read_workbook(path, clean)
        payload["validation"] = validate_payload(payload)
        return payload


def build_workbook(payload: dict[str, Any]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for key, (sheet_name, headers) in SHEET_SPECS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in payload.get(key, []):
            sheet.append([row.get(header, "") for header in headers])
    return workbook


def read_workbook(path: Path, scheme: str) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    payload: dict[str, Any] = {"scheme": scheme, "validation": [], "capacity_limits": []}
    for key, (sheet_name, headers) in SHEET_SPECS.items():
        if sheet_name not in workbook.sheetnames:
            payload[key] = []
            payload["validation"].append({"level": "error", "message": f"缺少工作表: {sheet_name}"})
            continue
        sheet = workbook[sheet_name]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if values is None or all(value is None for value in values):
                continue
            rows.append(
                {
                    header: values[index] if index < len(values) and values[index] is not None else ""
                    for index, header in enumerate(headers)
                }
            )
        payload[key] = rows
    return payload


def validate_payload(payload: dict[str, Any]) -> list[dict[str, str]]:
    messages: list[dict[str, str]] = []
    time_series = payload.get("time_series", [])
    if len(time_series) != 8760:
        messages.append({"level": "error", "message": f"8760时序数据行数应为8760，当前为{len(time_series)}"})
    else:
        messages.append({"level": "ok", "message": "8760时序数据行数正确"})

    for key in SHEET_SPECS:
        if key == "time_series":
            continue
        for index, row in enumerate(payload.get(key, []), start=1):
            lower = row.get("design_capacity_lower", "")
            upper = row.get("design_capacity_upper", "")
            if lower == "" or upper == "":
                continue
            try:
                lower_number = float(lower)
                upper_number = float(upper)
            except (TypeError, ValueError):
                messages.append({"level": "error", "message": f"{SHEET_SPECS[key][0]}第{index}行设计容量上下限不是数值"})
                continue
            if lower_number > upper_number:
                messages.append({"level": "error", "message": f"{SHEET_SPECS[key][0]}第{index}行设计容量上限不能小于下限"})
    return messages
